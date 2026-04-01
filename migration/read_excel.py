import openpyxl
import json

wb = openpyxl.load_workbook('2026-03-26T14_08_18.3898181Z.xlsx')

ws = wb['in']
print(f'Rows: {ws.max_row}, Cols: {ws.max_column}')

headers = [cell.value for cell in ws[1]]
for i, h in enumerate(headers):
    print(f'Col {i}: {h}')

print('\n--- ALL DATA ROWS ---')
all_rows = []
for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True)):
    all_rows.append(list(row))

# Write to JSON for easier reading
with open('excel_data.json', 'w', encoding='utf-8') as f:
    json.dump({'headers': headers, 'rows': all_rows}, f, indent=2, default=str)

print(f'Wrote {len(all_rows)} rows to excel_data.json')
